import io
import os
import tempfile
from pathlib import Path

TEST_DATA = tempfile.mkdtemp(prefix='acct_review_test_')
os.environ['ACCOUNTING_DATA_DIR'] = TEST_DATA

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.db import db, init_db, get_settings, set_settings
from app.main import app
from app.services import (
    detect_mapping, import_transactions, dashboard_summary, validate_all,
    checklist_data, monthly_report, resolve_issue, ai_assist_prompt,
)
from app.documents import save_document, UPLOAD_DIR
from app.exports import create_review_workbook, create_audit_package, create_sample_template

init_db()
client = TestClient(app)


def reset_data():
    with db() as conn:
        conn.execute('DELETE FROM issues')
        conn.execute('DELETE FROM transactions')
        conn.execute('DELETE FROM imports')
        conn.execute('DELETE FROM documents')
        conn.execute('DELETE FROM audit_log')
    for p in UPLOAD_DIR.glob('*'):
        if p.is_file():
            p.unlink()


def make_xlsx(rows=None):
    wb = Workbook()
    ws = wb.active
    ws.append(['전표번호','지출일자','금액','공급가액','부가세','거래처','사용목적','계정과목','부서','사용자','증빙번호','증빙여부','결제수단'])
    default = [
        ['A-001','2026-08-01',110000,100000,10000,'가나다상사','고객 미팅 식대','복리후생비','영업팀','김대리','R-001','있음','법인카드'],
        ['A-002','2026-08-02',220000,200000,20000,'테스트상사','장비 소모품 구입','소모품비','기술팀','이사원','R-002','없음','법인카드'],
        ['A-003','2026-08-03',330000,300000,20000,'오류상사','출장 교통비','여비교통비','서비스팀','박과장','','없음','법인카드'],
        ['A-004','2026-08-01',110000,100000,10000,'가나다상사','고객 미팅 식대','복리후생비','영업팀','최사원','R-004','있음','법인카드'],
    ]
    for r in (rows or default):
        ws.append(r)
    bio = io.BytesIO(); wb.save(bio); return bio.getvalue()


def setup_function(_):
    reset_data()
    set_settings({'high_amount_threshold':'1000000','purpose_min_length':'4','vat_tolerance':'1','evidence_required':'1'})


def test_mapping_detection():
    m = detect_mapping(['지출일자','금액','사용목적','거래처','증빙여부'])
    assert m['expense_date'] == '지출일자'
    assert m['amount'] == '금액'
    assert m['purpose'] == '사용목적'


def test_import_and_validation():
    result = import_transactions('sample.xlsx', make_xlsx())
    assert result['rows'] == 4
    s = dashboard_summary()
    assert s['total'] == 4
    assert s['review_required'] >= 3
    assert s['duplicates'] == 2
    assert s['missing_evidence'] >= 2
    with db() as conn:
        assert conn.execute("SELECT COUNT(*) FROM issues WHERE rule_code='VAT_MISMATCH'").fetchone()[0] == 1


def test_duplicate_file_rejected():
    data = make_xlsx()
    import_transactions('sample.xlsx', data)
    try:
        import_transactions('sample-copy.xlsx', data)
        assert False, 'duplicate file should fail'
    except ValueError as e:
        assert '이미 업로드' in str(e)


def test_issue_resolution_survives_revalidation():
    import_transactions('sample.xlsx', make_xlsx())
    with db() as conn:
        issue = conn.execute("SELECT id, transaction_id, rule_code FROM issues WHERE rule_code='VAT_MISMATCH' LIMIT 1").fetchone()
    resolve_issue(issue['id'], '예외인정', '원증빙 확인 후 예외 인정')
    validate_all()
    with db() as conn:
        row = conn.execute("SELECT status, resolution_note FROM issues WHERE transaction_id=? AND rule_code=?", (issue['transaction_id'], issue['rule_code'])).fetchone()
    assert row['status'] == '예외인정'
    assert '원증빙' in row['resolution_note']


def test_document_evidence_link_removes_missing_issue():
    rows = [['B-001','2026-08-04',50000,45455,4545,'문서상사','회의 문구 구입','소모품비','관리팀','정사원','DOC-77','없음','법인카드']]
    import_transactions('doc.xlsx', make_xlsx(rows))
    with db() as conn:
        assert conn.execute("SELECT COUNT(*) FROM issues WHERE rule_code='EVIDENCE_MISSING'").fetchone()[0] == 1
    save_document('receipt.txt', '영수증 승인번호 1234 공급가액 45,455'.encode('utf-8'), reference_no='DOC-77')
    with db() as conn:
        assert conn.execute("SELECT COUNT(*) FROM issues WHERE rule_code='EVIDENCE_MISSING'").fetchone()[0] == 0
        doc = conn.execute("SELECT category FROM documents LIMIT 1").fetchone()
    assert doc['category'] == '영수증'


def test_checklist_and_monthly_report():
    import_transactions('sample.xlsx', make_xlsx())
    ck = checklist_data()
    assert ck['total'] == 4
    assert not ck['ready']
    r = monthly_report('2026-08')
    assert r['count'] == 4
    assert r['amount'] == 770000
    assert '회계자료는 총 4건' in r['draft']


def test_ai_prompt_uses_verified_aggregates_only():
    import_transactions('sample.xlsx', make_xlsx())
    p = ai_assist_prompt('2026-08')['prompt']
    assert '770,000원' in p
    assert '숫자를 변경' in p
    assert '없는 원인' in p
    assert '김대리' not in p


def test_exports_created():
    import_transactions('sample.xlsx', make_xlsx())
    sample = create_sample_template()
    review = create_review_workbook('2026-08')
    package = create_audit_package('2026-08')
    assert sample.exists() and sample.stat().st_size > 1000
    assert review.exists() and review.stat().st_size > 1000
    assert package.exists() and package.stat().st_size > 1000


def test_settings_are_codeless():
    set_settings({'program_name':'테스트 회계 검토','menu_reports':'월간 보고서'})
    s = get_settings()
    assert s['program_name'] == '테스트 회계 검토'
    assert s['menu_reports'] == '월간 보고서'


def test_web_pages_render():
    import_transactions('sample.xlsx', make_xlsx())
    for path in ['/', '/upload', '/transactions', '/documents', '/checklist', '/reports', '/ai-assist', '/settings', '/health']:
        res = client.get(path)
        assert res.status_code == 200, (path, res.status_code, res.text[:500])


def test_mapping_preview_and_manual_blank_override():
    data = make_xlsx()
    files = {'file': ('preview.xlsx', data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
    res = client.post('/upload/preview', files=files)
    assert res.status_code == 200
    assert '열 매핑 확인' in res.text


def test_audit_log_written():
    import_transactions('sample.xlsx', make_xlsx())
    with db() as conn:
        actions = [r[0] for r in conn.execute('SELECT action FROM audit_log')]
    assert '자료 업로드' in actions
    assert '검토 기준 실행' in actions

def test_monthly_report_compares_previous_month():
    rows = [
        ['J-001','2026-07-10',100000,90909,9091,'전월상사','전월 회의비','회의비','관리팀','A','JUL-1','있음','법인카드'],
        ['J-002','2026-08-10',150000,136364,13636,'당월상사','당월 회의비','회의비','관리팀','B','AUG-1','있음','법인카드'],
    ]
    import_transactions('compare.xlsx', make_xlsx(rows))
    r = monthly_report('2026-08')
    assert r['prev_month'] == '2026-07'
    assert r['prev_amount'] == 100000
    assert round(r['amount_change_pct'], 1) == 50.0
    assert '+50.0%' in r['draft']


def test_month_export_filters_other_months():
    rows = [
        ['M-001','2026-07-01',10000,9091,909,'칠월상사','7월 비용','소모품비','관리팀','A','J1','있음','법인카드'],
        ['M-002','2026-08-01',20000,18182,1818,'팔월상사','8월 비용','소모품비','관리팀','B','A1','있음','법인카드'],
    ]
    import_transactions('months.xlsx', make_xlsx(rows))
    path = create_review_workbook('2026-08')
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb['회계자료']
    assert ws.max_row == 2
    assert ws['E2'].value == '2026-08-01'

def test_full_backup_created():
    from app.exports import create_full_backup
    import_transactions('sample.xlsx', make_xlsx())
    path = create_full_backup()
    assert path.exists() and path.stat().st_size > 1000


def test_v02_ui_text_is_codeless_and_reflected():
    with db() as conn:
        conn.execute("UPDATE ui_texts SET value='내가 바꾼 대시보드 안내' WHERE key='dashboard_description'")
    res = client.get('/')
    assert res.status_code == 200
    assert '내가 바꾼 대시보드 안내' in res.text


def test_v02_menu_order_visibility_and_label():
    with db() as conn:
        conn.execute("UPDATE menus SET label='맞춤 월간보고', sort_order=5, enabled=1 WHERE menu_key='reports'")
        conn.execute("UPDATE menus SET enabled=0 WHERE menu_key='ai'")
    res = client.get('/')
    assert '맞춤 월간보고' in res.text
    # hidden from nav, but the route itself remains accessible for safe recovery/admin use
    assert '>보고서 작성 보조</a>' not in res.text
    assert client.get('/ai-assist').status_code == 200


def test_v02_admin_can_add_custom_rule_and_revalidate():
    rows = [['C-001','2026-08-05',250000,227273,22727,'커스텀상사','테스트 구입','소모품비','관리팀','A','C-1','있음','법인카드']]
    import_transactions('custom.xlsx', make_xlsx(rows))
    res = client.post('/settings/rule-add', data={
        'name':'20만원 이상 확인','field_name':'amount','operator':'gte','compare_value':'200000',
        'severity':'확인','category':'회사규정','message':'20만원 이상 거래입니다. 내부승인 근거를 확인하세요.'
    }, follow_redirects=False)
    assert res.status_code == 303
    with db() as conn:
        rule = conn.execute("SELECT code FROM review_rules WHERE name='20만원 이상 확인'").fetchone()
        assert rule
        issue = conn.execute("SELECT * FROM issues WHERE rule_code=?", (rule['code'],)).fetchone()
        assert issue and issue['category'] == '회사규정'


def test_v02_builtin_rule_threshold_edit():
    rows = [['H-001','2026-08-05',300000,272727,27273,'기준상사','기준 테스트','소모품비','관리팀','A','H-1','있음','법인카드']]
    import_transactions('threshold.xlsx', make_xlsx(rows))
    with db() as conn:
        assert conn.execute("SELECT COUNT(*) FROM issues WHERE rule_code='AMOUNT_HIGH'").fetchone()[0] == 0
        r = conn.execute("SELECT * FROM review_rules WHERE code='AMOUNT_HIGH'").fetchone()
    res = client.post('/settings/rule/AMOUNT_HIGH', data={
        'name':r['name'],'enabled':'1','severity':r['severity'],'category':r['category'],
        'compare_value':'200000','message':r['message'],'sort_order':r['sort_order']
    }, follow_redirects=False)
    assert res.status_code == 303
    with db() as conn:
        assert conn.execute("SELECT COUNT(*) FROM issues WHERE rule_code='AMOUNT_HIGH'").fetchone()[0] == 1


def test_v02_issue_workflow_assignee_due_date_and_audit():
    rows = [['W-001','2026-08-05',50000,45455,4545,'워크상사','업무 지출','소모품비','관리팀','A','','없음','법인카드']]
    import_transactions('workflow.xlsx', make_xlsx(rows))
    with db() as conn:
        issue = conn.execute("SELECT id,transaction_id FROM issues WHERE rule_code='EVIDENCE_MISSING' LIMIT 1").fetchone()
    resolve_issue(issue['id'], '보완요청', '영수증 재요청', '김회계', '2026-08-20')
    with db() as conn:
        row = conn.execute("SELECT status,assignee,due_date,resolution_note FROM issues WHERE id=?", (issue['id'],)).fetchone()
        assert row['status'] == '보완요청' and row['assignee'] == '김회계' and row['due_date'] == '2026-08-20'
        assert conn.execute("SELECT COUNT(*) FROM audit_log WHERE action='검토 항목 처리'").fetchone()[0] >= 1
    resolve_issue(issue['id'], '확인완료', '증빙 수령', '김회계', '2026-08-20')
    with db() as conn:
        assert conn.execute("SELECT status FROM issues WHERE id=?", (issue['id'],)).fetchone()['status'] == '확인완료'


def test_v02_manual_checklist_confirmation():
    import_transactions('sample.xlsx', make_xlsx())
    with db() as conn:
        item = conn.execute("SELECT id FROM checklist_items WHERE item_type='MANUAL' ORDER BY id LIMIT 1").fetchone()
    from app.services import set_checklist_confirmation
    before = checklist_data()
    manual_before = [x for x in before['items'] if x['id'] == item['id']][0]
    assert manual_before['ok'] is False
    set_checklist_confirmation(item['id'], '전체', True, '원본 및 예외근거 확인')
    after = checklist_data()
    manual_after = [x for x in after['items'] if x['id'] == item['id']][0]
    assert manual_after['ok'] is True
    assert '원본' in manual_after['note']


def test_v02_monthly_report_has_observed_change_drivers():
    rows = [
        ['D-001','2026-07-10',100000,90909,9091,'A상사','7월 회의','회의비','관리팀','A','D1','있음','법인카드'],
        ['D-002','2026-08-10',300000,272727,27273,'A상사','8월 회의','회의비','관리팀','A','D2','있음','법인카드'],
        ['D-003','2026-08-11',50000,45455,4545,'B상사','8월 소모품','소모품비','관리팀','A','D3','있음','법인카드'],
    ]
    import_transactions('drivers.xlsx', make_xlsx(rows))
    r = monthly_report('2026-08')
    assert r['account_changes']
    assert r['account_changes'][0]['name'] == '회의비'
    assert r['account_changes'][0]['delta'] == 200000
    assert '원인을 의미하지 않습니다' in r['draft']


def test_v02_settings_page_contains_codeless_management_sections():
    res = client.get('/settings')
    assert res.status_code == 200
    for text in ['화면 문구 관리','검토 기준 관리','제출 체크리스트 관리','메뉴 이름·순서·표시 여부']:
        assert text in res.text


def test_v02_issue_work_queue_renders_and_dashboard_links_to_actions():
    import_transactions('sample.xlsx', make_xlsx())
    res = client.get('/issues')
    assert res.status_code == 200
    assert '확인·보완 업무' in res.text
    assert 'EVIDENCE_MISSING' in res.text
    dash = client.get('/')
    assert '/issues?rule=EVIDENCE_MISSING' in dash.text
    assert '/issues?assignee=__UNASSIGNED__' in dash.text


def test_v02_issue_work_queue_inline_update():
    rows = [['Q-001','2026-08-05',50000,45455,4545,'큐상사','업무 지출','소모품비','관리팀','A','','없음','법인카드']]
    import_transactions('queue.xlsx', make_xlsx(rows))
    with db() as conn:
        issue_id = conn.execute("SELECT id FROM issues WHERE rule_code='EVIDENCE_MISSING' LIMIT 1").fetchone()[0]
    res = client.post(f'/issues/{issue_id}/update', data={
        'status':'담당자지정','assignee':'최회계','due_date':'2026-08-25','note':'증빙 요청 예정'
    }, follow_redirects=False)
    assert res.status_code == 303
    with db() as conn:
        r = conn.execute('SELECT status,assignee,due_date FROM issues WHERE id=?',(issue_id,)).fetchone()
        assert (r['status'],r['assignee'],r['due_date']) == ('담당자지정','최회계','2026-08-25')


def test_v02_confirmation_message_template_is_codeless():
    rows = [['T-001','2026-08-05',50000,45455,4545,'템플릿상사','업무 지출','소모품비','관리팀','홍길동','','없음','법인카드']]
    import_transactions('template.xlsx', make_xlsx(rows))
    with db() as conn:
        conn.execute("UPDATE ui_texts SET value='[{date}] {vendor} {amount} 증빙을 보내주세요.' WHERE key='confirm_evidence'")
        txid = conn.execute("SELECT id FROM transactions LIMIT 1").fetchone()[0]
    tx = __import__('app.services', fromlist=['get_transaction_detail']).get_transaction_detail(txid)
    issue = next(i for i in tx['issues'] if i['rule_code']=='EVIDENCE_MISSING')
    msg = __import__('app.services', fromlist=['confirmation_message']).confirmation_message(tx, issue)
    assert '[2026-08-05] 템플릿상사 50,000원 증빙을 보내주세요.' in msg


def test_v02_mapping_profile_is_saved_and_reused():
    import re
    data1 = make_xlsx([['P-001','2026-08-01',10000,9091,909,'A','프로필 테스트','소모품비','관리','A','P1','있음','카드']])
    res = client.post('/upload/preview', files={'file':('profile1.xlsx',data1,'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')})
    assert res.status_code == 200
    token = re.search(r'name="token" value="([^"]+)"', res.text).group(1)
    mapping = {
        'token':token,'transaction_id':'전표번호','expense_date':'지출일자','amount':'금액','supply_amount':'공급가액','tax_amount':'부가세',
        'vendor':'거래처','purpose':'사용목적','account_name':'계정과목','department':'부서','employee':'사용자','evidence_no':'증빙번호',
        'evidence_status':'증빙여부','payment_method':'결제수단','note':''
    }
    res = client.post('/upload/confirm', data=mapping, follow_redirects=False)
    assert res.status_code == 303
    with db() as conn:
        assert conn.execute('SELECT COUNT(*) FROM mapping_profiles').fetchone()[0] == 1
    data2 = make_xlsx([['P-002','2026-08-02',20000,18182,1818,'B','프로필 재사용','소모품비','관리','B','P2','있음','카드']])
    res2 = client.post('/upload/preview', files={'file':('profile2.xlsx',data2,'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')})
    assert res2.status_code == 200
    assert '저장된 매핑 프로필 적용' in res2.text
    assert 'value="지출일자" selected' in res2.text


def test_v02_invalid_numeric_rule_setting_is_rejected():
    with db() as conn:
        r = conn.execute("SELECT * FROM review_rules WHERE code='AMOUNT_HIGH'").fetchone()
        old = r['compare_value']
    res = client.post('/settings/rule/AMOUNT_HIGH', data={
        'name':r['name'],'enabled':'1','severity':r['severity'],'category':r['category'],
        'compare_value':'not-a-number','message':r['message'],'sort_order':r['sort_order']
    }, follow_redirects=False)
    assert res.status_code == 400
    with db() as conn:
        assert conn.execute("SELECT compare_value FROM review_rules WHERE code='AMOUNT_HIGH'").fetchone()[0] == old


def test_v02_checklist_rejects_unknown_rule_code():
    res = client.post('/settings/checklist-add', data={
        'label':'잘못된 자동항목','item_type':'AUTO_RULES','rule_codes':'NOT_EXIST_RULE','severity_filter':''
    }, follow_redirects=False)
    assert res.status_code == 400
    with db() as conn:
        assert conn.execute("SELECT COUNT(*) FROM checklist_items WHERE label='잘못된 자동항목'").fetchone()[0] == 0
