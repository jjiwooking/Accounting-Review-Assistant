from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from statistics import median
from typing import Any, Iterable

from openpyxl import load_workbook

from .db import db, get_settings, get_rules, get_ui_texts
from .rules_engine import rule_matches, render_message

COLUMN_SYNONYMS: dict[str, list[str]] = {
    "transaction_id": ["거래번호", "전표번호", "전표 no", "전표no", "transaction id", "id", "번호"],
    "expense_date": ["지출일자", "사용일자", "거래일자", "일자", "날짜", "date", "expense date"],
    "amount": ["금액", "합계", "총액", "결제금액", "지출금액", "amount", "total"],
    "supply_amount": ["공급가액", "공급금액", "과세표준", "supply amount", "net amount"],
    "tax_amount": ["부가세", "세액", "vat", "tax", "tax amount"],
    "vendor": ["거래처", "가맹점", "업체명", "공급자", "vendor", "merchant", "supplier"],
    "purpose": ["사용목적", "지출목적", "적요", "내용", "용도", "purpose", "description", "memo"],
    "account_name": ["계정과목", "계정", "비용계정", "account", "account name"],
    "department": ["부서", "소속부서", "코스트센터", "cost center", "department"],
    "employee": ["사용자", "직원", "성명", "신청자", "사용자명", "employee", "user"],
    "evidence_no": ["증빙번호", "증빙 no", "증빙no", "영수증번호", "세금계산서번호", "evidence no", "receipt no"],
    "evidence_status": ["증빙여부", "증빙", "영수증여부", "첨부여부", "evidence", "receipt attached"],
    "payment_method": ["결제수단", "지급수단", "카드/현금", "payment method", "method"],
    "note": ["비고", "메모", "note", "remarks"],
}

DISPLAY_NAMES = {
    "transaction_id": "거래/전표번호",
    "expense_date": "지출일자",
    "amount": "금액",
    "supply_amount": "공급가액",
    "tax_amount": "부가세",
    "vendor": "거래처",
    "purpose": "사용목적",
    "account_name": "계정과목",
    "department": "부서",
    "employee": "사용자",
    "evidence_no": "증빙번호",
    "evidence_status": "증빙여부",
    "payment_method": "결제수단",
    "note": "비고",
}

MAX_ROWS = 100_000

TRUE_WORDS = {"1", "y", "yes", "true", "o", "있음", "첨부", "첨부됨", "완료", "확인", "확인됨", "제출"}
FALSE_WORDS = {"0", "n", "no", "false", "x", "없음", "미첨부", "미제출", "누락"}


def now_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat(sep=" ")


def norm_header(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", " ", text.strip().lower())
    return text


def detect_mapping(headers: list[str]) -> dict[str, str]:
    normalized = {norm_header(h): h for h in headers if h is not None}
    mapping: dict[str, str] = {}
    for target, synonyms in COLUMN_SYNONYMS.items():
        for synonym in synonyms:
            s = norm_header(synonym)
            if s in normalized:
                mapping[target] = normalized[s]
                break
        if target in mapping:
            continue
        # relaxed matching only when the synonym is meaningful enough
        for nh, original in normalized.items():
            if any(len(norm_header(syn)) >= 3 and norm_header(syn) in nh for syn in synonyms):
                mapping[target] = original
                break
    return mapping


def parse_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d", "%y-%m-%d", "%y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    # Excel serial date fallback is intentionally not guessed from arbitrary numeric data.
    return None


def parse_amount(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.replace(",", "").replace("₩", "").replace("원", "").replace(" ", "")
    text = text.replace("(", "").replace(")", "")
    try:
        result = float(text)
        return -result if negative else result
    except ValueError:
        return None


def normalize_evidence(value: Any) -> str | None:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip().lower()
    if text in TRUE_WORDS:
        return "있음"
    if text in FALSE_WORDS:
        return "없음"
    return str(value).strip()


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def read_table_bytes(filename: str, content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        # UTF-8 BOM first, then cp949 fallback common in Korean Excel exports.
        decoded = None
        for enc in ("utf-8-sig", "cp949", "euc-kr"):
            try:
                decoded = content.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if decoded is None:
            raise ValueError("CSV 문자 인코딩을 읽을 수 없습니다. UTF-8 또는 CP949로 저장해 주세요.")
        reader = csv.DictReader(io.StringIO(decoded))
        headers = [h or "" for h in (reader.fieldnames or [])]
        rows = []
        for i, r in enumerate(reader):
            if i >= MAX_ROWS:
                rows.append({"__ROW_LIMIT__": True})
                break
            rows.append(dict(r))
        return headers, rows
    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        values = ws.iter_rows(values_only=True)
        try:
            header_row = next(values)
        except StopIteration:
            return [], []
        headers = ["" if v is None else str(v).strip() for v in header_row]
        rows: list[dict[str, Any]] = []
        for row_values in values:
            if not any(v not in (None, "") for v in row_values):
                continue
            if len(rows) >= MAX_ROWS:
                rows.append({"__ROW_LIMIT__": True})
                break
            rows.append({headers[i]: row_values[i] if i < len(row_values) else None for i in range(len(headers))})
        return headers, rows
    raise ValueError("지원 형식은 CSV, XLSX, XLSM입니다.")


def standardize_row(row: dict[str, Any], mapping: dict[str, str]) -> dict[str, Any]:
    def raw(key: str) -> Any:
        source = mapping.get(key)
        return row.get(source) if source else None

    return {
        "transaction_id": clean_text(raw("transaction_id")),
        "expense_date": parse_date(raw("expense_date")),
        "amount": parse_amount(raw("amount")),
        "supply_amount": parse_amount(raw("supply_amount")),
        "tax_amount": parse_amount(raw("tax_amount")),
        "vendor": clean_text(raw("vendor")),
        "purpose": clean_text(raw("purpose")),
        "account_name": clean_text(raw("account_name")),
        "department": clean_text(raw("department")),
        "employee": clean_text(raw("employee")),
        "evidence_no": clean_text(raw("evidence_no")),
        "evidence_status": normalize_evidence(raw("evidence_status")),
        "payment_method": clean_text(raw("payment_method")),
        "note": clean_text(raw("note")),
    }


def import_transactions(filename: str, content: bytes, mapping_override: dict[str, str] | None = None) -> dict[str, Any]:
    headers, rows = read_table_bytes(filename, content)
    if not headers:
        raise ValueError("헤더가 없는 파일입니다.")
    if len(rows) > MAX_ROWS or (rows and rows[-1].get("__ROW_LIMIT__")):
        raise ValueError(f"한 번에 최대 {MAX_ROWS:,}행까지 처리할 수 있습니다. 파일을 기간별로 나눠 업로드해 주세요.")
    mapping = detect_mapping(headers)
    if mapping_override is not None:
        for key, value in mapping_override.items():
            if value:
                mapping[key] = value
            else:
                mapping.pop(key, None)
    required = {"expense_date", "amount", "purpose"}
    missing = [DISPLAY_NAMES[k] for k in required if k not in mapping]
    if missing:
        raise ValueError("필수 열을 자동 인식하지 못했습니다: " + ", ".join(missing))

    timestamp = now_iso()
    file_hash = hashlib.sha256(content).hexdigest()
    with db() as conn:
        existing = conn.execute("SELECT id, filename FROM imports WHERE file_hash=?", (file_hash,)).fetchone()
        if existing:
            raise ValueError(f"동일한 파일이 이미 업로드되어 있습니다: {existing['filename']}")
        cur = conn.execute(
            "INSERT INTO imports(filename, file_hash, imported_at, row_count, mapped_columns) VALUES(?, ?, ?, ?, ?)",
            (filename, file_hash, timestamp, len(rows), json.dumps(mapping, ensure_ascii=False)),
        )
        import_id = cur.lastrowid
        inserted = 0
        for idx, row in enumerate(rows, start=2):
            s = standardize_row(row, mapping)
            conn.execute(
                """
                INSERT INTO transactions(
                    import_id, source_file, source_row, transaction_id, expense_date, amount,
                    supply_amount, tax_amount, vendor, purpose, account_name, department,
                    employee, evidence_no, evidence_status, payment_method, note, created_at
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    import_id, filename, idx, s["transaction_id"], s["expense_date"], s["amount"],
                    s["supply_amount"], s["tax_amount"], s["vendor"], s["purpose"], s["account_name"],
                    s["department"], s["employee"], s["evidence_no"], s["evidence_status"],
                    s["payment_method"], s["note"], timestamp,
                ),
            )
            inserted += 1
        conn.execute(
            "INSERT INTO audit_log(action, target_type, target_id, detail, created_at) VALUES(?,?,?,?,?)",
            ("자료 업로드", "import", str(import_id), f"{filename} / {inserted}건", timestamp),
        )
    validate_all()
    return {"import_id": import_id, "rows": inserted, "mapping": mapping, "headers": headers}


@dataclass
class Finding:
    rule_code: str
    severity: str
    category: str
    message: str


def _document_refs(conn) -> set[str]:
    return {str(r[0]).strip().lower() for r in conn.execute("SELECT reference_no FROM documents WHERE reference_no IS NOT NULL AND TRIM(reference_no) <> ''")}


def _basic_findings(tx: dict[str, Any], rules: list[dict[str, Any]], evidence_refs: set[str]) -> list[Finding]:
    findings: list[Finding] = []
    for rule in rules:
        if rule.get("rule_type") in {"DUPLICATE_EXACT", "DUPLICATE_PROBABLE"}:
            continue
        if rule_matches(tx, rule, evidence_refs):
            findings.append(Finding(
                rule["code"], rule["severity"], rule["category"], render_message(tx, rule)
            ))
    return findings


def validate_all() -> dict[str, int]:
    rules = get_rules(include_disabled=False)
    rule_by_type = {r["rule_type"]: r for r in rules}
    created = 0
    with db() as conn:
        evidence_refs = _document_refs(conn)
        rows = [dict(r) for r in conn.execute("SELECT * FROM transactions ORDER BY id")]
        preserved = {
            (r["transaction_id"], r["rule_code"]): (
                r["status"], r["resolution_note"], r["resolved_at"], r["assignee"], r["due_date"]
            )
            for r in conn.execute("SELECT transaction_id, rule_code, status, resolution_note, resolved_at, assignee, due_date FROM issues")
        }
        conn.execute("DELETE FROM issues")

        by_id: dict[int, list[Finding]] = {}
        for tx in rows:
            by_id[tx["id"]] = _basic_findings(tx, rules, evidence_refs)

        # Duplicate rules remain deterministic but their use/severity/message are admin controlled.
        exact_rule = rule_by_type.get("DUPLICATE_EXACT")
        probable_rule = rule_by_type.get("DUPLICATE_PROBABLE")
        exact_groups: dict[tuple[Any, ...], list[int]] = defaultdict(list)
        probable_groups: dict[tuple[Any, ...], list[int]] = defaultdict(list)
        for tx in rows:
            if tx.get("expense_date") and tx.get("amount") is not None:
                vendor = (tx.get("vendor") or "").strip().lower()
                purpose = (tx.get("purpose") or "").strip().lower()
                exact_groups[(tx["expense_date"], round(float(tx["amount"]), 2), vendor, purpose)].append(tx["id"])
                probable_groups[(tx["expense_date"], round(float(tx["amount"]), 2), vendor)].append(tx["id"])

        exact_ids: set[int] = set()
        if exact_rule:
            for ids in exact_groups.values():
                if len(ids) > 1:
                    exact_ids.update(ids)
                    for tx_id in ids:
                        msg = exact_rule["message"] + f" (동일 조건 {len(ids)}건)"
                        by_id[tx_id].append(Finding(exact_rule["code"], exact_rule["severity"], exact_rule["category"], msg))
        if probable_rule:
            for ids in probable_groups.values():
                if len(ids) > 1:
                    for tx_id in ids:
                        if tx_id not in exact_ids:
                            msg = probable_rule["message"] + f" (유사 조건 {len(ids)}건)"
                            by_id[tx_id].append(Finding(probable_rule["code"], probable_rule["severity"], probable_rule["category"], msg))

        timestamp = now_iso()
        for tx in rows:
            findings = by_id[tx["id"]]
            open_for_tx = 0
            for item in findings:
                old_status, old_note, old_resolved, old_assignee, old_due = preserved.get(
                    (tx["id"], item.rule_code), ("미확인", None, None, None, None)
                )
                conn.execute(
                    """INSERT INTO issues(
                        transaction_id, rule_code, severity, category, message, status, resolution_note,
                        created_at, resolved_at, assignee, due_date
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                    (tx["id"], item.rule_code, item.severity, item.category, item.message,
                     old_status, old_note, timestamp, old_resolved, old_assignee, old_due),
                )
                if old_status not in {"확인완료", "예외인정"}:
                    open_for_tx += 1
                created += 1
            status = "정상" if open_for_tx == 0 else "검토필요"
            conn.execute("UPDATE transactions SET review_status=? WHERE id=?", (status, tx["id"]))
        conn.execute(
            "INSERT INTO audit_log(action, target_type, detail, created_at) VALUES(?,?,?,?)",
            ("검토 기준 실행", "validation", f"{len(rows)}건 / 검토 항목 {created}건 / 사용 기준 {len(rules)}개", timestamp),
        )
    return {"transactions": len(rows), "issues": created, "rules": len(rules)}

def dashboard_summary() -> dict[str, Any]:
    terminal = ("확인완료", "예외인정")
    with db() as conn:
        row = conn.execute(
            """
            SELECT COUNT(*) total,
                   COALESCE(SUM(amount),0) total_amount,
                   SUM(CASE WHEN review_status='검토필요' THEN 1 ELSE 0 END) review_required,
                   SUM(CASE WHEN review_status='정상' THEN 1 ELSE 0 END) normal_count
            FROM transactions
            """
        ).fetchone()
        issue_counts = {r["severity"]: r["cnt"] for r in conn.execute(
            "SELECT severity, COUNT(*) cnt FROM issues WHERE status NOT IN ('확인완료','예외인정') GROUP BY severity"
        )}
        missing_evidence = conn.execute(
            "SELECT COUNT(DISTINCT transaction_id) FROM issues WHERE rule_code='EVIDENCE_MISSING' AND status NOT IN ('확인완료','예외인정')"
        ).fetchone()[0]
        duplicates = conn.execute(
            "SELECT COUNT(DISTINCT transaction_id) FROM issues WHERE rule_code LIKE 'DUPLICATE_%' AND status NOT IN ('확인완료','예외인정')"
        ).fetchone()[0]
        unassigned = conn.execute(
            "SELECT COUNT(*) FROM issues WHERE status NOT IN ('확인완료','예외인정') AND (assignee IS NULL OR TRIM(assignee)='')"
        ).fetchone()[0]
        overdue = conn.execute(
            "SELECT COUNT(*) FROM issues WHERE status NOT IN ('확인완료','예외인정') AND due_date IS NOT NULL AND due_date < date('now','localtime')"
        ).fetchone()[0]
        imports = conn.execute("SELECT COUNT(*) FROM imports").fetchone()[0]
        return {
            "total": row["total"] or 0,
            "total_amount": row["total_amount"] or 0,
            "review_required": row["review_required"] or 0,
            "normal_count": row["normal_count"] or 0,
            "errors": issue_counts.get("오류", 0),
            "warnings": issue_counts.get("주의", 0),
            "checks": issue_counts.get("확인", 0),
            "missing_evidence": missing_evidence,
            "duplicates": duplicates,
            "unassigned": unassigned,
            "overdue": overdue,
            "imports": imports,
        }

def get_transactions(q: str = "", status: str = "", month: str = "") -> list[dict[str, Any]]:
    clauses: list[str] = []
    params: list[Any] = []
    if q:
        clauses.append("(COALESCE(vendor,'') LIKE ? OR COALESCE(purpose,'') LIKE ? OR COALESCE(transaction_id,'') LIKE ? OR COALESCE(employee,'') LIKE ?)")
        like = f"%{q}%"
        params.extend([like, like, like, like])
    if status:
        clauses.append("review_status=?")
        params.append(status)
    if month:
        clauses.append("SUBSTR(expense_date,1,7)=?")
        params.append(month)
    where = "WHERE " + " AND ".join(clauses) if clauses else ""
    with db() as conn:
        rows = [dict(r) for r in conn.execute(f"SELECT * FROM transactions {where} ORDER BY expense_date DESC, id DESC", params)]
        for row in rows:
            row["issue_count"] = conn.execute("SELECT COUNT(*) FROM issues WHERE transaction_id=? AND status NOT IN ('확인완료','예외인정')", (row["id"],)).fetchone()[0]
        return rows


def get_transaction_detail(tx_id: int) -> dict[str, Any] | None:
    with db() as conn:
        row = conn.execute("SELECT * FROM transactions WHERE id=?", (tx_id,)).fetchone()
        if not row:
            return None
        result = dict(row)
        result["issues"] = [dict(r) for r in conn.execute("SELECT * FROM issues WHERE transaction_id=? ORDER BY CASE severity WHEN '오류' THEN 1 WHEN '주의' THEN 2 ELSE 3 END, id", (tx_id,))]
        result["matched_documents"] = [dict(r) for r in conn.execute("SELECT * FROM documents WHERE reference_no=? ORDER BY id DESC", (result.get("evidence_no"),))] if result.get("evidence_no") else []
        return result


def resolve_issue(issue_id: int, status: str, note: str = "", assignee: str = "", due_date: str = "") -> None:
    allowed = {"미확인", "담당자지정", "보완요청", "재검토", "확인완료", "예외인정"}
    if status not in allowed:
        raise ValueError("올바르지 않은 처리상태입니다.")
    with db() as conn:
        issue = conn.execute("SELECT transaction_id, status, assignee, due_date FROM issues WHERE id=?", (issue_id,)).fetchone()
        if not issue:
            raise ValueError("검토 항목을 찾을 수 없습니다.")
        resolved_at = now_iso() if status in {"확인완료", "예외인정"} else None
        assignee_v = assignee.strip() or None
        due_v = due_date.strip() or None
        conn.execute(
            "UPDATE issues SET status=?, resolution_note=?, resolved_at=?, assignee=?, due_date=? WHERE id=?",
            (status, note.strip() or None, resolved_at, assignee_v, due_v, issue_id),
        )
        remaining = conn.execute(
            "SELECT COUNT(*) FROM issues WHERE transaction_id=? AND status NOT IN ('확인완료','예외인정')",
            (issue["transaction_id"],),
        ).fetchone()[0]
        conn.execute("UPDATE transactions SET review_status=? WHERE id=?", ("정상" if remaining == 0 else "검토필요", issue["transaction_id"]))
        detail = f"{issue['status']}→{status} / 담당:{assignee_v or '-'} / 기한:{due_v or '-'} / {note}"
        conn.execute(
            "INSERT INTO audit_log(action,target_type,target_id,detail,created_at) VALUES(?,?,?,?,?)",
            ("검토 항목 처리", "issue", str(issue_id), detail[:500], now_iso()),
        )

def update_transaction(tx_id: int, reviewer_note: str, review_status: str | None = None) -> None:
    with db() as conn:
        if review_status and review_status not in {"미검토", "검토필요", "정상"}:
            raise ValueError("검토상태가 올바르지 않습니다.")
        if review_status:
            conn.execute("UPDATE transactions SET reviewer_note=?, review_status=? WHERE id=?", (reviewer_note.strip() or None, review_status, tx_id))
        else:
            conn.execute("UPDATE transactions SET reviewer_note=? WHERE id=?", (reviewer_note.strip() or None, tx_id))
        conn.execute("INSERT INTO audit_log(action,target_type,target_id,detail,created_at) VALUES(?,?,?,?,?)", ("검토 메모", "transaction", str(tx_id), reviewer_note[:200], now_iso()))


def checklist_data(month: str | None = None) -> dict[str, Any]:
    period = month or "전체"
    with db() as conn:
        if month:
            tx_ids = [r[0] for r in conn.execute("SELECT id FROM transactions WHERE SUBSTR(expense_date,1,7)=?", (month,))]
        else:
            tx_ids = [r[0] for r in conn.execute("SELECT id FROM transactions")]
        total = len(tx_ids)
        categories: list[dict[str, Any]] = []
        unresolved = unresolved_tx = errors = warnings = 0
        if tx_ids:
            ph = ",".join("?" for _ in tx_ids)
            open_where = f"transaction_id IN ({ph}) AND status NOT IN ('확인완료','예외인정')"
            unresolved = conn.execute(f"SELECT COUNT(*) FROM issues WHERE {open_where}", tx_ids).fetchone()[0]
            unresolved_tx = conn.execute(f"SELECT COUNT(DISTINCT transaction_id) FROM issues WHERE {open_where}", tx_ids).fetchone()[0]
            errors = conn.execute(f"SELECT COUNT(*) FROM issues WHERE {open_where} AND severity='오류'", tx_ids).fetchone()[0]
            warnings = conn.execute(f"SELECT COUNT(*) FROM issues WHERE {open_where} AND severity='주의'", tx_ids).fetchone()[0]
            categories = [dict(r) for r in conn.execute(
                f"SELECT category, COUNT(*) count FROM issues WHERE {open_where} GROUP BY category ORDER BY count DESC", tx_ids
            )]
        else:
            open_where = "1=0"
            ph = ""

        items = []
        for item in conn.execute("SELECT * FROM checklist_items WHERE enabled=1 ORDER BY sort_order,id"):
            d = dict(item)
            if d["item_type"] == "MANUAL":
                conf = conn.execute(
                    "SELECT checked,note FROM checklist_confirmations WHERE item_id=? AND period=?", (d["id"], period)
                ).fetchone()
                d["ok"] = bool(conf and conf["checked"])
                d["note"] = conf["note"] if conf else ""
            elif not tx_ids:
                d["ok"] = False
                d["note"] = "검토할 자료가 없습니다."
            elif d["item_type"] == "AUTO_SEVERITY":
                cnt = conn.execute(
                    f"SELECT COUNT(*) FROM issues WHERE {open_where} AND severity=?", tx_ids + [d["severity_filter"]]
                ).fetchone()[0]
                d["ok"] = cnt == 0
                d["note"] = f"미완료 {cnt}건"
            else:  # AUTO_RULES
                codes = [x.strip() for x in (d.get("rule_codes") or "").split(",") if x.strip()]
                if codes:
                    q = ",".join("?" for _ in codes)
                    cnt = conn.execute(
                        f"SELECT COUNT(*) FROM issues WHERE {open_where} AND rule_code IN ({q})", tx_ids + codes
                    ).fetchone()[0]
                else:
                    cnt = 0
                d["ok"] = cnt == 0
                d["note"] = f"미완료 {cnt}건"
            items.append(d)
        ready = total > 0 and bool(items) and all(i["ok"] for i in items)
        return {
            "month": period, "total": total, "unresolved": unresolved, "unresolved_tx": unresolved_tx,
            "errors": errors, "warnings": warnings, "categories": categories, "items": items, "ready": ready,
        }


def set_checklist_confirmation(item_id: int, period: str, checked: bool, note: str = "") -> None:
    with db() as conn:
        conn.execute(
            """INSERT INTO checklist_confirmations(item_id,period,checked,note,updated_at)
               VALUES(?,?,?,?,?) ON CONFLICT(item_id,period) DO UPDATE SET checked=excluded.checked,note=excluded.note,updated_at=excluded.updated_at""",
            (item_id, period or "전체", 1 if checked else 0, note.strip() or None, now_iso()),
        )
        conn.execute(
            "INSERT INTO audit_log(action,target_type,target_id,detail,created_at) VALUES(?,?,?,?,?)",
            ("체크리스트 확인", "checklist", str(item_id), f"{period or '전체'} / {'완료' if checked else '해제'} / {note}", now_iso()),
        )

def confirmation_message(tx: dict[str, Any], issue: dict[str, Any]) -> str:
    texts = get_ui_texts()
    values = {
        "name": tx.get("employee") or "담당자",
        "date": tx.get("expense_date") or "일자 미확인",
        "vendor": tx.get("vendor") or "해당 거래처",
        "amount": f"{tx.get('amount'):,.0f}원" if tx.get("amount") is not None else "금액 미확인",
        "message": issue.get("message", ""),
    }
    rule = issue.get("rule_code", "")
    if rule == "EVIDENCE_MISSING":
        key = "confirm_evidence"
    elif rule.startswith("DUPLICATE_"):
        key = "confirm_duplicate"
    elif rule in {"PURPOSE_MISSING", "PURPOSE_SHORT"}:
        key = "confirm_purpose"
    elif rule == "VAT_MISMATCH":
        key = "confirm_vat"
    else:
        key = "confirm_default"
    def safe_format(template: str) -> str:
        try:
            return template.format(**values)
        except (KeyError, ValueError):
            return template
    greeting = safe_format(texts.get("confirm_greeting", "안녕하세요, {name}님."))
    body = safe_format(texts.get(key, "{message}"))
    closing = safe_format(texts.get("confirm_closing", "확인 감사합니다."))
    return f"{greeting}\n\n{body}\n\n{closing}"


def months_available() -> list[str]:
    with db() as conn:
        return [r[0] for r in conn.execute("SELECT DISTINCT SUBSTR(expense_date,1,7) m FROM transactions WHERE expense_date IS NOT NULL ORDER BY m DESC")]


def monthly_report(month: str | None = None) -> dict[str, Any]:
    months = months_available()
    if not month:
        month = months[0] if months else ""
    with db() as conn:
        if not month:
            return {"month": "", "count": 0, "amount": 0, "issue_count": 0, "draft": "분석할 데이터가 없습니다.", "categories": [], "vendors": [], "months": months}
        txs = [dict(r) for r in conn.execute("SELECT * FROM transactions WHERE SUBSTR(expense_date,1,7)=?", (month,))]
        ids = [t["id"] for t in txs]
        amount = sum(float(t["amount"] or 0) for t in txs)
        issue_count = 0
        error_count = 0
        evidence_count = 0
        duplicate_count = 0
        if ids:
            placeholders = ",".join("?" for _ in ids)
            issue_count = conn.execute(f"SELECT COUNT(*) FROM issues WHERE transaction_id IN ({placeholders}) AND status NOT IN ('확인완료','예외인정')", ids).fetchone()[0]
            error_count = conn.execute(f"SELECT COUNT(*) FROM issues WHERE transaction_id IN ({placeholders}) AND status NOT IN ('확인완료','예외인정') AND severity='오류'", ids).fetchone()[0]
            evidence_count = conn.execute(f"SELECT COUNT(*) FROM issues WHERE transaction_id IN ({placeholders}) AND status NOT IN ('확인완료','예외인정') AND rule_code='EVIDENCE_MISSING'", ids).fetchone()[0]
            duplicate_count = conn.execute(f"SELECT COUNT(*) FROM issues WHERE transaction_id IN ({placeholders}) AND status NOT IN ('확인완료','예외인정') AND rule_code LIKE 'DUPLICATE_%'", ids).fetchone()[0]
        cat_counter = Counter((t.get("account_name") or "미분류") for t in txs)
        cat_amount = defaultdict(float)
        vendor_amount = defaultdict(float)
        for t in txs:
            cat_amount[t.get("account_name") or "미분류"] += float(t.get("amount") or 0)
            vendor_amount[t.get("vendor") or "거래처 미입력"] += float(t.get("amount") or 0)
        categories = sorted(({"name": k, "amount": v, "count": cat_counter[k]} for k, v in cat_amount.items()), key=lambda x: x["amount"], reverse=True)[:8]
        vendors = sorted(({"name": k, "amount": v} for k, v in vendor_amount.items()), key=lambda x: x["amount"], reverse=True)[:8]

        y, m = map(int, month.split("-"))
        prev_y, prev_m = (y - 1, 12) if m == 1 else (y, m - 1)
        prev_month = f"{prev_y:04d}-{prev_m:02d}"
        prev_row = conn.execute("SELECT COUNT(*) cnt, COALESCE(SUM(amount),0) amt FROM transactions WHERE SUBSTR(expense_date,1,7)=?", (prev_month,)).fetchone()
        prev_count = prev_row["cnt"] or 0
        prev_amount = float(prev_row["amt"] or 0)
        amount_change_pct = ((amount - prev_amount) / abs(prev_amount) * 100) if prev_amount else None

        # Explain change through observed account/vendor deltas. This is descriptive, not causal.
        prev_txs = [dict(r) for r in conn.execute("SELECT * FROM transactions WHERE SUBSTR(expense_date,1,7)=?", (prev_month,))]
        prev_cat = defaultdict(float); prev_vendor = defaultdict(float)
        for t in prev_txs:
            prev_cat[t.get("account_name") or "미분류"] += float(t.get("amount") or 0)
            prev_vendor[t.get("vendor") or "거래처 미입력"] += float(t.get("amount") or 0)
        account_keys = set(cat_amount) | set(prev_cat)
        vendor_keys = set(vendor_amount) | set(prev_vendor)
        account_changes = sorted(
            ({"name": k, "current": cat_amount.get(k,0.0), "previous": prev_cat.get(k,0.0), "delta": cat_amount.get(k,0.0)-prev_cat.get(k,0.0)} for k in account_keys),
            key=lambda x: abs(x["delta"]), reverse=True
        )[:5]
        vendor_changes = sorted(
            ({"name": k, "current": vendor_amount.get(k,0.0), "previous": prev_vendor.get(k,0.0), "delta": vendor_amount.get(k,0.0)-prev_vendor.get(k,0.0)} for k in vendor_keys),
            key=lambda x: abs(x["delta"]), reverse=True
        )[:5]
        driver_text = ""
        if prev_count and account_changes:
            top = account_changes[0]
            direction = "증가" if top["delta"] >= 0 else "감소"
            driver_text = f" 계정과목 기준 가장 큰 변동은 {top['name']} {abs(top['delta']):,.0f}원 {direction}입니다. 이는 관측된 금액 변화이며 원인을 의미하지 않습니다."
        ready_text = "제출 전 검토가 가능한 상태입니다." if error_count == 0 else "오류 등급 미확인 항목이 남아 있어 제출 전 추가 검토가 필요합니다."
        compare_text = (f" 전월({prev_month}) 대비 금액은 {amount_change_pct:+.1f}% 변동했습니다." if amount_change_pct is not None else " 비교 가능한 전월 금액 데이터는 없습니다.")
        draft = (
            f"{month} 회계자료는 총 {len(txs):,}건, 합계 {amount:,.0f}원입니다."
            f"{compare_text} 현재 미처리 검토 항목은 {issue_count:,}건이며, 이 중 오류 등급은 {error_count:,}건입니다. "
            f"증빙 누락 확인 건은 {evidence_count:,}건, 중복·분할결제 의심 건은 {duplicate_count:,}건입니다. "
            f"{driver_text} {ready_text}"
        )
        return {
            "month": month, "months": months, "count": len(txs), "amount": amount,
            "prev_month": prev_month, "prev_count": prev_count, "prev_amount": prev_amount, "amount_change_pct": amount_change_pct,
            "issue_count": issue_count, "error_count": error_count, "evidence_count": evidence_count,
            "duplicate_count": duplicate_count, "categories": categories, "vendors": vendors, "account_changes": account_changes, "vendor_changes": vendor_changes, "draft": draft,
        }


def data_quality_score() -> dict[str, Any]:
    summary = dashboard_summary()
    total = summary["total"]
    if total == 0:
        return {"score": 0, "label": "자료 없음", "note": "자료를 등록하면 점검 현황을 계산합니다."}
    with db() as conn:
        critical_tx = conn.execute("SELECT COUNT(DISTINCT transaction_id) FROM issues WHERE severity='오류' AND status NOT IN ('확인완료','예외인정')").fetchone()[0]
        warning_tx = conn.execute("SELECT COUNT(DISTINCT transaction_id) FROM issues WHERE severity='주의' AND status NOT IN ('확인완료','예외인정')").fetchone()[0]
    score = max(0, round(100 - (critical_tx / total * 70) - (warning_tx / total * 20)))
    label = "양호" if score >= 90 else "검토 필요" if score >= 70 else "주의"
    note = "이 점수는 제출 적정성을 판단하는 법적 기준이 아니라, 프로그램의 검토 기준에 따른 내부 점검 지표입니다."
    return {"score": score, "label": label, "note": note}


def ai_assist_prompt(month: str | None = None) -> dict[str, str]:
    report = monthly_report(month)
    if not report.get("month"):
        return {"month": "", "prompt": "분석할 데이터가 없습니다."}
    category_lines = "\n".join(
        f"- {c['name']}: {c['amount']:,.0f}원 / {c['count']}건" for c in report.get("categories", [])[:5]
    ) or "- 계정과목 데이터 없음"
    prompt = f"""당신은 회계자료 월간 보고 문장 작성 보조자입니다. 아래에 제공된 검증 완료 수치만 사용하세요.

[절대 규칙]
1. 아래 숫자를 변경하거나 새 숫자를 계산·추정하지 마세요.
2. 없는 원인·사실·규정 위반을 만들어내지 마세요.
3. '오류', '증빙누락', '중복의심'은 프로그램의 검토 기준에 따른 확인 대상이며 부정·위법을 의미하지 않습니다.
4. 회계·세무 판단이 추가로 필요한 부분은 '담당자 확인 필요'라고 명시하세요.
5. 한국 기업의 내부 월간 회계보고 문체로, 간결하게 작성하세요.

[검증된 사실]
- 대상월: {report['month']}
- 회계자료: {report['count']:,}건
- 합계금액: {report['amount']:,.0f}원
- 미처리 검토 항목: {report['issue_count']:,}건
- 오류 등급: {report.get('error_count',0):,}건
- 증빙 누락 확인: {report.get('evidence_count',0):,}건
- 중복·분할결제 의심: {report.get('duplicate_count',0):,}건

[금액 상위 계정과목]
{category_lines}

[전월 대비 관측된 주요 변동]
{chr(10).join(f"- {x['name']}: {x['delta']:+,.0f}원" for x in report.get('account_changes', [])[:3]) or '- 비교 가능한 전월 데이터 없음'}
※ 위 변동은 원인 추정이 아니라 실제 집계 차이입니다. 원인은 증빙·계약·부서 확인이 필요합니다.

[요청 결과]
A. 경영진용 5줄 요약
B. 회계담당자가 제출 전 확인할 항목 3개
C. 과장 없이 중립적인 월간 보고서 초안
"""
    return {"month": report["month"], "prompt": prompt}
