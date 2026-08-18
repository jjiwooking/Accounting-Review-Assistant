from __future__ import annotations

import io
import json
import shutil
import sqlite3
import tempfile
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .db import BASE_DIR, DB_PATH, db, get_settings, get_rules, get_menus, list_ui_texts
from .services import checklist_data, monthly_report, now_iso

EXPORT_DIR = BASE_DIR / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

HEADERS_FILL = PatternFill("solid", fgColor="D9EAF7")
ERROR_FILL = PatternFill("solid", fgColor="FCE8E6")
WARN_FILL = PatternFill("solid", fgColor="FFF4CE")


def _style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADERS_FILL
        cell.alignment = Alignment(horizontal="center")
    for col in range(1, ws.max_column + 1):
        max_len = 10
        for row in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row, col).value
            if value is not None:
                max_len = max(max_len, min(50, len(str(value)) + 2))
        ws.column_dimensions[get_column_letter(col)].width = max_len


def create_sample_template() -> Path:
    path = EXPORT_DIR / "회계자료_업로드_표준양식.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "회계자료"
    headers = ["전표번호", "지출일자", "금액", "공급가액", "부가세", "거래처", "사용목적", "계정과목", "부서", "사용자", "증빙번호", "증빙여부", "결제수단", "비고"]
    ws.append(headers)
    ws.append(["EXP-001", "2026-08-01", 110000, 100000, 10000, "예시상사", "고객 미팅 식대", "복리후생비", "영업팀", "홍길동", "R-001", "있음", "법인카드", "예시 행 - 삭제 후 사용"])
    _style_sheet(ws)
    wb.save(path)
    return path



def create_demo_data() -> Path:
    path = EXPORT_DIR / "회계자료_검토_데모.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "회계자료"
    ws.append(["전표번호","지출일자","금액","공급가액","부가세","거래처","사용목적","계정과목","부서","사용자","증빙번호","증빙여부","결제수단","비고"])
    rows = [
        ["D-001","2026-07-15",88000,80000,8000,"서울문구","사무용품 구입","소모품비","경영지원팀","김사원","RC-001","있음","법인카드","정상 예시"],
        ["D-002","2026-08-03",165000,150000,15000,"비즈식당","고객 미팅 식대","접대비","영업팀","이대리","RC-002","없음","법인카드","증빙 누락 예시"],
        ["D-003","2026-08-05",330000,300000,20000,"KTX","출장 교통비","여비교통비","서비스팀","박과장","","없음","법인카드","부가세 합계 불일치 예시"],
        ["D-004","2026-08-10",1200000,1090909,109091,"ABC솔루션","연간 소프트웨어 사용료","지급수수료","IT팀","최대리","RC-004","있음","법인카드","고액 확인 예시"],
        ["D-005","2026-08-12",55000,50000,5000,"카페테스트","회의","회의비","개발팀","정사원","RC-005","있음","법인카드","사용목적 짧음"],
        ["D-006","2026-08-15",99000,90000,9000,"택배상사","고객 샘플 발송","운반비","영업팀","홍사원","RC-006","있음","법인카드","중복 예시 1"],
        ["D-007","2026-08-15",99000,90000,9000,"택배상사","고객 샘플 발송","운반비","영업팀","홍사원","RC-007","있음","법인카드","중복 예시 2"],
    ]
    for r in rows: ws.append(r)
    _style_sheet(ws); wb.save(path); return path

def create_review_workbook(month: str | None = None) -> Path:
    settings = get_settings()
    with db() as conn:
        if month:
            txs = [dict(r) for r in conn.execute("SELECT * FROM transactions WHERE SUBSTR(expense_date,1,7)=? ORDER BY expense_date, id", (month,))]
            issues = [dict(r) for r in conn.execute("""
                SELECT i.*, t.source_file, t.source_row, t.expense_date, t.amount, t.vendor, t.purpose, t.evidence_no
                FROM issues i JOIN transactions t ON t.id=i.transaction_id
                WHERE SUBSTR(t.expense_date,1,7)=?
                ORDER BY CASE i.severity WHEN '오류' THEN 1 WHEN '주의' THEN 2 ELSE 3 END, i.id
            """, (month,))]
            refs = sorted({str(t.get("evidence_no")).strip() for t in txs if t.get("evidence_no")})
            if refs:
                ph = ",".join("?" for _ in refs)
                docs = [dict(r) for r in conn.execute(f"SELECT * FROM documents WHERE reference_no IN ({ph}) ORDER BY id", refs)]
            else:
                docs = []
        else:
            txs = [dict(r) for r in conn.execute("SELECT * FROM transactions ORDER BY expense_date, id")]
            issues = [dict(r) for r in conn.execute("""
                SELECT i.*, t.source_file, t.source_row, t.expense_date, t.amount, t.vendor, t.purpose, t.evidence_no
                FROM issues i JOIN transactions t ON t.id=i.transaction_id
                ORDER BY CASE i.severity WHEN '오류' THEN 1 WHEN '주의' THEN 2 ELSE 3 END, i.id
            """)]
            docs = [dict(r) for r in conn.execute("SELECT * FROM documents ORDER BY id")]
        audits = [dict(r) for r in conn.execute("SELECT * FROM audit_log ORDER BY id DESC LIMIT 1000")]
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("회계자료")
    headers = ["ID", "원본파일", "원본행", "전표번호", "지출일자", "금액", "공급가액", "부가세", "거래처", "사용목적", "계정과목", "부서", "사용자", "증빙번호", "증빙여부", "결제수단", "검토상태", "검토메모"]
    ws.append(headers)
    for t in txs:
        ws.append([t["id"], t["source_file"], t["source_row"], t["transaction_id"], t["expense_date"], t["amount"], t["supply_amount"], t["tax_amount"], t["vendor"], t["purpose"], t["account_name"], t["department"], t["employee"], t["evidence_no"], t["evidence_status"], t["payment_method"], t["review_status"], t["reviewer_note"]])
    _style_sheet(ws)

    ws = wb.create_sheet("검토이슈")
    headers = ["이슈ID", "거래ID", "등급", "분류", "규칙", "메시지", "처리상태", "담당자", "처리기한", "처리메모", "원본파일", "원본행", "지출일자", "금액", "거래처", "사용목적", "증빙번호"]
    ws.append(headers)
    for i in issues:
        ws.append([i["id"], i["transaction_id"], i["severity"], i["category"], i["rule_code"], i["message"], i["status"], i.get("assignee"), i.get("due_date"), i["resolution_note"], i["source_file"], i["source_row"], i["expense_date"], i["amount"], i["vendor"], i["purpose"], i["evidence_no"]])
        if i["severity"] == "오류":
            for cell in ws[ws.max_row]:
                cell.fill = ERROR_FILL
        elif i["severity"] == "주의":
            for cell in ws[ws.max_row]:
                cell.fill = WARN_FILL
    _style_sheet(ws)

    ck = checklist_data(month)
    ws = wb.create_sheet("제출체크리스트")
    ws.append(["항목", "상태"])
    for item in ck["items"]:
        ws.append([item["label"], "완료" if item["ok"] else "확인필요"])
    ws.append([])
    ws.append(["전체 자료건수", ck["total"]])
    ws.append(["미완료 이슈", ck["unresolved"]])
    ws.append(["제출 준비 상태", "준비됨" if ck["ready"] else "추가 검토 필요"])
    _style_sheet(ws)

    report = monthly_report(month)
    ws = wb.create_sheet("월별보고")
    ws.append(["항목", "내용"])
    ws.append(["대상월", report["month"]])
    ws.append(["자료건수", report["count"]])
    ws.append(["합계금액", report["amount"]])
    ws.append(["미완료 이슈", report["issue_count"]])
    ws.append(["오류 등급", report.get("error_count", 0)])
    ws.append(["증빙 누락", report.get("evidence_count", 0)])
    ws.append(["중복 의심", report.get("duplicate_count", 0)])
    ws.append(["보고 초안", report["draft"]])
    _style_sheet(ws)

    ws = wb.create_sheet("전월증감")
    ws.append(["구분", "항목", "전월", "당월", "증감"])
    for x in report.get("account_changes", []):
        ws.append(["계정과목", x["name"], x["previous"], x["current"], x["delta"]])
    for x in report.get("vendor_changes", []):
        ws.append(["거래처", x["name"], x["previous"], x["current"], x["delta"]])
    _style_sheet(ws)

    ws = wb.create_sheet("검토Rule설정")
    ws.append(["코드", "이름", "사용", "유형", "필드", "조건", "기준값", "등급", "분류", "메시지", "순서", "기본Rule"])
    for r in get_rules(include_disabled=True):
        ws.append([r["code"], r["name"], "사용" if r["enabled"] else "미사용", r["rule_type"], r.get("field_name"), r.get("operator"), r.get("compare_value"), r["severity"], r["category"], r["message"], r["sort_order"], "기본" if r["is_system"] else "사용자"])
    _style_sheet(ws)

    ws = wb.create_sheet("화면메뉴설정")
    ws.append(["구분", "키", "값", "표시/순서"])
    for m in get_menus(include_disabled=True):
        ws.append(["메뉴", m["menu_key"], m["label"], f"{'표시' if m['enabled'] else '숨김'} / {m['sort_order']}"])
    for x in list_ui_texts():
        ws.append(["문구", x["key"], x["value"], x["group_name"]])
    _style_sheet(ws)

    ws = wb.create_sheet("증빙문서")
    ws.append(["문서ID", "파일명", "분류", "증빙번호", "요약", "업로드일"])
    for d in docs:
        ws.append([d["id"], d["filename"], d["category"], d["reference_no"], d["summary"], d["uploaded_at"]])
    _style_sheet(ws)

    ws = wb.create_sheet("변경이력")
    ws.append(["일시", "작업", "대상", "대상ID", "상세"])
    for a in audits:
        ws.append([a["created_at"], a["action"], a["target_type"], a["target_id"], a["detail"]])
    _style_sheet(ws)

    safe_month = report["month"].replace("-", "") if report["month"] else "전체"
    path = EXPORT_DIR / f"회계자료_검토보고_{safe_month}.xlsx"
    wb.save(path)
    return path


def create_audit_package(month: str | None = None) -> Path:
    review_xlsx = create_review_workbook(month)
    settings = get_settings()
    with db() as conn:
        if month:
            refs = [r[0] for r in conn.execute("SELECT DISTINCT evidence_no FROM transactions WHERE SUBSTR(expense_date,1,7)=? AND evidence_no IS NOT NULL AND TRIM(evidence_no)<>''", (month,))]
            if refs:
                ph = ",".join("?" for _ in refs)
                docs = [dict(r) for r in conn.execute(f"SELECT * FROM documents WHERE reference_no IN ({ph}) ORDER BY id", refs)]
            else:
                docs = []
        else:
            docs = [dict(r) for r in conn.execute("SELECT * FROM documents ORDER BY id")]
    stamp = now_iso().replace(":", "").replace("-", "").replace(" ", "_")
    path = EXPORT_DIR / f"감사대응_검토패키지_{stamp}.zip"
    manifest = {
        "created_at": now_iso(),
        "company": settings.get("report_company_name"),
        "period": month or "전체",
        "note": "프로그램 자동 검토 결과는 내부 검토 보조자료이며 회계·세무·감사인의 전문적 판단을 대체하지 않습니다.",
        "document_count": len(docs),
    }
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(review_xlsx, arcname=f"01_검토보고/{review_xlsx.name}")
        zf.writestr("00_안내/manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
        for d in docs:
            src = BASE_DIR / "uploads" / d["stored_name"]
            if src.exists():
                ref = (d.get("reference_no") or "미지정").replace("/", "_")
                zf.write(src, arcname=f"02_증빙문서/{ref}_{d['filename']}")
    return path


def create_full_backup() -> Path:
    stamp = now_iso().replace(":", "").replace("-", "").replace(" ", "_")
    path = EXPORT_DIR / f"회계검토프로그램_전체백업_{stamp}.zip"
    with tempfile.TemporaryDirectory() as td:
        db_copy = Path(td) / "accounting_review.db"
        src = sqlite3.connect(DB_PATH)
        dst = sqlite3.connect(db_copy)
        try:
            src.backup(dst)
        finally:
            dst.close(); src.close()
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.write(db_copy, arcname="data/accounting_review.db")
            upload_dir = BASE_DIR / "uploads"
            for f in upload_dir.glob("*"):
                if f.is_file():
                    zf.write(f, arcname=f"uploads/{f.name}")
            zf.writestr("README_BACKUP.txt", "프로그램을 종료한 뒤 data/accounting_review.db와 uploads 폴더를 같은 위치에 복원하면 됩니다. 백업본은 개인정보·회계정보를 포함할 수 있으므로 안전하게 보관하세요.")
    return path
